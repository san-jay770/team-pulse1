"""
TEAM PULSE — Automated Excel Report & Email Delivery Module
Generates beautifully styled Excel workbooks (.xlsx) with openpyxl
and delivers them directly to Super Admin email inboxes.
"""

import os
import smtplib
import sqlite3
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load environment variables from .env file
def _load_env_file():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    env_paths = [
        os.path.join(base_dir, ".env"),
        os.path.join(os.path.dirname(base_dir), ".env")
    ]
    for env_path in env_paths:
        if os.path.exists(env_path):
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        k, v = line.split("=", 1)
                        k = k.strip()
                        v = v.strip().strip("'\"")
                        if k and not os.environ.get(k):
                            os.environ[k] = v

_load_env_file()

EMAIL_HOST = os.environ.get("EMAIL_HOST", os.environ.get("SMTP_HOST", "smtp.gmail.com"))
EMAIL_PORT = int(os.environ.get("EMAIL_PORT", os.environ.get("SMTP_PORT", 587)))
EMAIL_USER = os.environ.get("EMAIL_USER", os.environ.get("SMTP_USER", ""))
EMAIL_PASSWORD = os.environ.get("EMAIL_PASSWORD", os.environ.get("SMTP_PASS", ""))
EMAIL_FROM = os.environ.get("EMAIL_FROM", os.environ.get("SMTP_FROM", EMAIL_USER or "notifications@teampulse.local"))



def get_db_connection(db_path=None):
    if not db_path:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(base_dir, "team_pulse.db")
        if not os.path.exists(db_path):
            db_path = os.path.join(os.path.dirname(base_dir), "team_pulse.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def generate_excel_report(db_path=None):
    """
    Creates a styled multi-sheet Excel workbook containing:
    1. Executive Summary
    2. Team Performance
    3. Tasks List
    4. Member Leaderboard
    """
    conn = get_db_connection(db_path)
    wb = Workbook()

    # Define Styles
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    title_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")

    sub_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    sub_font = Font(name="Calibri", size=10, italic=True, color="E2E8F0")

    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    data_font = Font(name="Calibri", size=10, color="0F172A")
    bold_data_font = Font(name="Calibri", size=10, bold=True, color="0F172A")

    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    now_str = datetime.now().strftime("%d %B %Y, %I:%M %p")

    # ----------------------------------------------------
    # SHEET 1: Executive Summary
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner
    ws1.merge_cells("A1:D1")
    cell_title = ws1["A1"]
    cell_title.value = "TEAM PULSE — EXECUTIVE SYSTEM REPORT"
    cell_title.fill = title_fill
    cell_title.font = title_font
    cell_title.alignment = center_align
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:D2")
    cell_sub = ws1["A2"]
    cell_sub.value = f"Generated on: {now_str} | Automated System Snapshot"
    cell_sub.fill = sub_fill
    cell_sub.font = sub_font
    cell_sub.alignment = center_align
    ws1.row_dimensions[2].height = 20

    ws1.append([])

    # Metrics Gathering
    total_teams = conn.execute("SELECT COUNT(*) c FROM teams").fetchone()["c"]
    total_members = conn.execute("SELECT COUNT(*) c FROM users WHERE role='TEAM_MEMBER'").fetchone()["c"]
    total_admins = conn.execute("SELECT COUNT(*) c FROM users WHERE role IN ('ADMIN', 'SUPER_ADMIN')").fetchone()["c"]
    total_tasks = conn.execute("SELECT COUNT(*) c FROM tasks").fetchone()["c"]
    completed_tasks = conn.execute("SELECT COUNT(*) c FROM tasks WHERE status='Completed'").fetchone()["c"]
    pending_tasks = conn.execute("SELECT COUNT(*) c FROM tasks WHERE status!='Completed'").fetchone()["c"]
    overdue_tasks = conn.execute("SELECT COUNT(*) c FROM tasks WHERE status!='Completed' AND due_date IS NOT NULL AND due_date < date('now')").fetchone()["c"]
    total_points = conn.execute("SELECT COALESCE(SUM(points), 0) c FROM users").fetchone()["c"]

    summary_headers = ["Metric Category", "Description", "Value", "Status"]
    ws1.append(summary_headers)
    header_row_num = ws1.max_row
    ws1.row_dimensions[header_row_num].height = 24

    for col in range(1, 5):
        cell = ws1.cell(row=header_row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    summary_data = [
        ("Total Workspaces", "Active teams configured in the platform", total_teams, "Active"),
        ("Team Members", "Registered member accounts", total_members, "Active"),
        ("Administrators", "Team Admins & Super Admins", total_admins, "Active"),
        ("Total Tasks", "All-time task volume", total_tasks, "Total"),
        ("Completed Tasks", "Tasks finished by team members", completed_tasks, "Success"),
        ("Pending Tasks", "Tasks currently open or in progress", pending_tasks, "Pending"),
        ("Overdue Tasks", "Tasks past assigned due date", overdue_tasks, "Attention" if overdue_tasks > 0 else "Good"),
        ("Total Points Earned", "Cumulative reward points across all members", f"{total_points} PTS", "Active")
    ]

    for idx, (cat, desc, val, st) in enumerate(summary_data, start=1):
        ws1.append([cat, desc, val, st])
        r = ws1.max_row
        ws1.row_dimensions[r].height = 20
        is_even = idx % 2 == 0
        for col_idx in range(1, 5):
            c = ws1.cell(row=r, column=col_idx)
            c.font = bold_data_font if col_idx == 3 else data_font
            c.border = thin_border
            if is_even:
                c.fill = zebra_fill
            if col_idx in [3, 4]:
                c.alignment = center_align
            else:
                c.alignment = left_align

    # ----------------------------------------------------
    # SHEET 2: Team Performance
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Team Performance")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:E1")
    ws2["A1"].value = "TEAM PERFORMANCE & COMPLETION METRICS"
    ws2["A1"].fill = title_fill
    ws2["A1"].font = title_font
    ws2["A1"].alignment = center_align
    ws2.row_dimensions[1].height = 30

    team_headers = ["Team Name", "Admin Lead", "Total Tasks", "Completed Tasks", "Completion Rate (%)"]
    ws2.append(team_headers)
    ws2.row_dimensions[2].height = 24

    for col in range(1, 6):
        cell = ws2.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    teams_rows = conn.execute("""
        SELECT
            t.name AS team_name,
            a.name AS admin_name,
            COUNT(tk.id) AS total_tasks,
            COALESCE(SUM(CASE WHEN tk.status='Completed' THEN 1 ELSE 0 END), 0) AS completed_tasks
        FROM teams t
        LEFT JOIN users a ON t.admin_id=a.id
        LEFT JOIN tasks tk ON tk.team_id=t.id
        GROUP BY t.id, t.name
        ORDER BY t.id
    """).fetchall()

    for idx, t in enumerate(teams_rows, start=1):
        tot = t["total_tasks"] or 0
        done = t["completed_tasks"] or 0
        pct = round((done / tot) * 100) if tot > 0 else 0
        ws2.append([t["team_name"], t["admin_name"] or "Unassigned", tot, done, f"{pct}%"])
        r = ws2.max_row
        ws2.row_dimensions[r].height = 20
        is_even = idx % 2 == 0
        for c_idx in range(1, 6):
            c = ws2.cell(row=r, column=c_idx)
            c.font = data_font
            c.border = thin_border
            if is_even:
                c.fill = zebra_fill
            if c_idx >= 3:
                c.alignment = center_align

    # ----------------------------------------------------
    # SHEET 3: Tasks Detail
    # ----------------------------------------------------
    ws3 = wb.create_sheet(title="Tasks Master List")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:G1")
    ws3["A1"].value = "MASTER TASKS & DELIVERABLES INVENTORY"
    ws3["A1"].fill = title_fill
    ws3["A1"].font = title_font
    ws3["A1"].alignment = center_align
    ws3.row_dimensions[1].height = 30

    task_headers = ["Task Title", "Team", "Assigned Member", "Priority", "Status", "Due Date", "Created Date"]
    ws3.append(task_headers)
    ws3.row_dimensions[2].height = 24

    for col in range(1, 8):
        cell = ws3.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    task_rows = conn.execute("""
        SELECT
            tk.title,
            t.name AS team_name,
            u.name AS assignee,
            tk.priority,
            tk.status,
            tk.due_date,
            tk.created_at
        FROM tasks tk
        LEFT JOIN teams t ON tk.team_id=t.id
        LEFT JOIN users u ON tk.assigned_to=u.id
        ORDER BY tk.created_at DESC
    """).fetchall()

    for idx, tk in enumerate(task_rows, start=1):
        ws3.append([
            tk["title"],
            tk["team_name"] or "Unassigned",
            tk["assignee"] or "Unassigned",
            tk["priority"] or "Medium",
            tk["status"] or "Pending",
            tk["due_date"] or "-",
            tk["created_at"][:10] if tk["created_at"] else "-"
        ])
        r = ws3.max_row
        ws3.row_dimensions[r].height = 20
        is_even = idx % 2 == 0
        for c_idx in range(1, 8):
            c = ws3.cell(row=r, column=c_idx)
            c.font = data_font
            c.border = thin_border
            if is_even:
                c.fill = zebra_fill
            if c_idx in [4, 5, 6, 7]:
                c.alignment = center_align

    # ----------------------------------------------------
    # SHEET 4: Member Rankings & Points
    # ----------------------------------------------------
    ws4 = wb.create_sheet(title="Member Leaderboard")
    ws4.views.sheetView[0].showGridLines = True

    ws4.merge_cells("A1:D1")
    ws4["A1"].value = "MEMBER PERFORMANCE RANKINGS & POINTS"
    ws4["A1"].fill = title_fill
    ws4["A1"].font = title_font
    ws4["A1"].alignment = center_align
    ws4.row_dimensions[1].height = 30

    leaderboard_headers = ["Rank", "Member Name", "Team Workspace", "Total Points"]
    ws4.append(leaderboard_headers)
    ws4.row_dimensions[2].height = 24

    for col in range(1, 5):
        cell = ws4.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    leader_rows = conn.execute("""
        SELECT
            u.name,
            t.name AS team_name,
            u.points
        FROM users u
        LEFT JOIN teams t ON u.team_id=t.id
        WHERE u.role='TEAM_MEMBER'
        ORDER BY u.points DESC, u.name ASC
    """).fetchall()

    for idx, mem in enumerate(leader_rows, start=1):
        ws4.append([f"#{idx}", mem["name"], mem["team_name"] or "Unassigned", f"{mem['points']} PTS"])
        r = ws4.max_row
        ws4.row_dimensions[r].height = 20
        is_even = idx % 2 == 0
        for c_idx in range(1, 5):
            c = ws4.cell(row=r, column=c_idx)
            c.font = bold_data_font if c_idx == 4 else data_font
            c.border = thin_border
            if is_even:
                c.fill = zebra_fill
            if c_idx in [1, 4]:
                c.alignment = center_align

    # Auto Column Width Adjuster
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    conn.close()
    return wb


def send_excel_report_email(db_path=None, target_emails=None):
    """
    Generates the Excel report and sends it via email with attachment to Super Admin(s).
    """
    conn = get_db_connection(db_path)

    if not target_emails:
        # Fetch all active Super Admins
        rows = conn.execute("SELECT DISTINCT email FROM users WHERE role='SUPER_ADMIN' AND status='Active'").fetchall()
        target_emails = [r["email"] for r in rows if r["email"]]
        
        # Also fallback to env setting
        config_email = os.environ.get("SUPER_ADMIN_EMAIL")
        if config_email and config_email not in target_emails:
            target_emails.append(config_email)

    conn.close()

    if not target_emails:
        print("[Excel Email] Warning: No recipient Super Admin email addresses found.")
        return {"success": False, "message": "No recipient Super Admin email addresses found."}

    # Generate Workbook
    wb = generate_excel_report(db_path)
    
    # Save to temp buffer
    today_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"Team_Pulse_Executive_Report_{today_str}.xlsx"
    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scratch")
    os.makedirs(temp_dir, exist_ok=True)
    filepath = os.path.join(temp_dir, filename)
    wb.save(filepath)

    # Read binary bytes for attachment
    with open(filepath, "rb") as f:
        excel_bytes = f.read()

    # Build Email Message
    subject = f"📊 Team Pulse — Automated Executive Excel Report ({datetime.now().strftime('%d %b %Y')})"
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
      <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; background-color: #f8fafc; color: #0f172a; margin: 0; padding: 20px; }}
        .card {{ max-width: 600px; margin: 0 auto; background: white; border-radius: 12px; padding: 28px; box-shadow: 0 4px 12px rgba(0,0,0,0.08); border: 1px solid #e2e8f0; }}
        .header {{ display: flex; align-items: center; gap: 12px; margin-bottom: 20px; border-bottom: 2px solid #2563eb; padding-bottom: 14px; }}
        .brand-icon {{ background: #2563eb; color: white; font-weight: 800; font-size: 18px; width: 36px; height: 36px; border-radius: 8px; text-align: center; line-height: 36px; }}
        .title {{ font-size: 20px; font-weight: 800; color: #0f172a; margin: 0; }}
        .badge {{ background: #dbeafe; color: #1e40af; font-size: 12px; font-weight: 700; padding: 4px 10px; border-radius: 12px; display: inline-block; margin-bottom: 16px; }}
        .btn {{ display: inline-block; background: #2563eb; color: white; font-weight: 700; text-decoration: none; padding: 12px 24px; border-radius: 8px; margin-top: 20px; }}
        .footer {{ font-size: 12px; color: #64748b; margin-top: 24px; text-align: center; border-top: 1px solid #e2e8f0; padding-top: 14px; }}
      </style>
    </head>
    <body>
      <div class="card">
        <div class="header">
          <div class="brand-icon">TP</div>
          <h2 class="title">TEAM PULSE EXECUTIVE REPORT</h2>
        </div>
        <span class="badge">AUTOMATED EXCEL ATTACHMENT</span>
        <p>Dear Super Admin,</p>
        <p>Your scheduled <strong>Team Pulse Executive Performance & Master Inventory Report</strong> has been generated and attached to this email as an formatted Excel workbook (<code>{filename}</code>).</p>
        
        <p><strong>The attached workbook includes:</strong></p>
        <ul>
          <li><strong>Executive Summary:</strong> Overview of active teams, members, total tasks & point balances.</li>
          <li><strong>Team Performance:</strong> Completion percentages and task volume breakdown per team workspace.</li>
          <li><strong>Tasks Master Inventory:</strong> Complete deliverables log with status, priority, and assignees.</li>
          <li><strong>Member Leaderboard:</strong> Live member rank standings & accumulated reward points.</li>
        </ul>

        <p style="margin-top: 20px;">You can also access live interactive analytics directly on the Team Pulse dashboard.</p>
        
        <div class="footer">
          Team Pulse — Automated Executive Reporting System<br>
          Generated on {datetime.now().strftime('%A, %d %B %Y at %I:%M %p')}
        </div>
      </div>
    </body>
    </html>
    """

    # Read live environment parameters
    _load_env_file()
    email_user = os.environ.get("EMAIL_USER", os.environ.get("SMTP_USER", ""))
    email_pass = os.environ.get("EMAIL_PASSWORD", os.environ.get("SMTP_PASS", ""))
    email_host = os.environ.get("EMAIL_HOST", os.environ.get("SMTP_HOST", "smtp.gmail.com"))
    email_port = int(os.environ.get("EMAIL_PORT", os.environ.get("SMTP_PORT", 587)))
    email_from = os.environ.get("EMAIL_FROM", os.environ.get("SMTP_FROM", email_user or "notifications@teampulse.local"))

    # Check if SMTP Credentials are configured
    if not email_user or not email_pass or email_user == "your-email@gmail.com":
        print(f"[Excel Email] (Simulation Mode) Excel report generated successfully: {filepath}")
        print(f"  Target Recipients: {', '.join(target_emails)}")
        print("  Note: Configure EMAIL_USER and EMAIL_PASSWORD in .env for live Gmail delivery.")
        return {
            "success": True,
            "simulated": True,
            "filename": filename,
            "filepath": filepath,
            "recipients": target_emails,
            "message": "Excel report generated! (Simulation mode - configure EMAIL_USER & EMAIL_PASSWORD for live sending)"
        }

    try:
        msg = MIMEMultipart()
        msg["From"] = f"Team Pulse System <{email_user}>"
        msg["To"] = ", ".join(target_emails)
        msg["Subject"] = subject

        msg.attach(MIMEText(html_body, "html"))

        # Attach Excel file
        part = MIMEApplication(excel_bytes, Name=filename)
        part["Content-Disposition"] = f'attachment; filename="{filename}"'
        msg.attach(part)

        # Connect & Send
        server = smtplib.SMTP(email_host, email_port)
        server.starttls()
        server.login(email_user, email_pass)
        server.sendmail(email_user, target_emails, msg.as_string())
        server.quit()


        print(f"[Excel Email Sent Successfully] Excel report '{filename}' sent to {', '.join(target_emails)}")
        return {
            "success": True,
            "filename": filename,
            "recipients": target_emails,
            "message": f"Excel report delivered successfully to {', '.join(target_emails)}!"
        }

    except Exception as err:
        print(f"[Excel Email Error] Failed to send report: {err}")
        return {"success": False, "error": str(err), "message": f"Failed to deliver email: {err}"}


if __name__ == "__main__":
    res = send_excel_report_email()
    print("Test Result:", res)
